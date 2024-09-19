from openpyxl import load_workbook
import sys
import os

def update_xlsx(file_path, filename):
    try:
        if not os.path.isfile(file_path):
            print(f"File does not exist: {file_path}")
            return

        print(f"Updating file: {file_path} with filename: {filename}")

        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Check if headers exist, and if not, add them
        headers = [cell.value for cell in sheet[1]]
        print(f"Headers found: {headers}")

        if 'FILENAME' not in headers:
            sheet.cell(row=1, column=len(headers)+1, value='FILENAME')
            sheet.cell(row=1, column=len(headers)+2, value='ACTION')
        else:
            # Adjust column index if FILENAME and ACTION already exist
            filename_col = headers.index('FILENAME') + 1
            action_col = headers.index('ACTION') + 1

        # Iterate over rows and add/update 'FILENAME' and 'ACTION'
        for row in sheet.iter_rows(min_row=2):
            if len(row) >= 2:  # Ensure row has at least two columns
                row[-2].value = filename  # Update FILENAME
                row[-1].value = 'empty'   # Update ACTION to 'empty'

        # Save the workbook with changes
        workbook.save(file_path)
        print(f"File updated successfully: {file_path}")
    
    except Exception as e:
        print(f"Error updating file: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python update_xlsx.py <file_path> <filename>")
        sys.exit(1)

    file_path = sys.argv[1]  # Path to the XLSX file
    filename = sys.argv[2]   # FILENAME to be added
    update_xlsx(file_path, filename)
