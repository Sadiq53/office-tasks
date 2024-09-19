from openpyxl import load_workbook
import sys

def update_xlsx(file_path, agreement_number, action_status):
    # Load the existing workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Determine column indices for ACTION and FILENAME
    header_row = sheet[1]
    action_col_idx = None

    # Find column indices for ACTION and FILENAME
    for col_idx, cell in enumerate(header_row, 1):
        if cell.value == 'ACTION':
            action_col_idx = col_idx
        elif cell.value is None and action_col_idx is None:
            action_col_idx = col_idx
            sheet.cell(row=1, column=action_col_idx, value='ACTION')

    # Update the ACTION column for the specific agreement_number
    for row in sheet.iter_rows(min_row=2):
        cell_value = row[0].value  # Assuming the agreement number is in the first column
        if cell_value == agreement_number:
            # Update ACTION field
            row[action_col_idx - 1].value = action_status
            break

    # Save the workbook with changes
    workbook.save(file_path)

if __name__ == "__main__":
    file_path = sys.argv[1]          # Path to the XLSX file
    agreement_number = sys.argv[2]   # Agreement number to find
    action_status = sys.argv[3]      # New action status to set
    update_xlsx(file_path, agreement_number, action_status)
